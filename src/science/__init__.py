import os
from io import BytesIO

from matplotlib.figure import Figure
from matplotlib.backends.backend_qt5 import FigureCanvasQT

from openpyxl import load_workbook

FACTORS = [
    "Без нагрузки",
    "С физической нагрузкой",
    "С эмоциональной нагрузкой",
    "После отдыха"
]
FACTORS_L = [factor.lower() for factor in FACTORS]
FACTORS_ALL = -1


class ScienceError(Exception):
    pass


class ClassesError(ValueError):
    pass


class ParseError(ValueError):
    pass

#достаем имя файла без типа из path
def file_base_name(filename: str):
    return os.path.splitext(os.path.basename(filename))[0]


def read_xlsx_std(filename):
    try:
        #для чтения xlsx
        wb = load_workbook(filename=filename)
    except Exception:
        raise ParseError("Невозможно открыть файл образца {}"
                         "\nВозможно он открыт в другой программе".format(filename))
    #инициализация листа с информацией о пациенте из xlsx
    sheet = wb.get_active_sheet()
    datas = []
    rows = sheet.max_row

    for i in range(2, rows + 1):
        data = []
        for j in range(1, 19 + 1):
            try:
                cell = sheet.cell(row=i, column=j)
                if sheet.cell(1, j).value == "date":
                    data.append(str(cell.value))
                else:
                    data.append(float(cell.value))
            except Exception:
                raise ParseError('Ошибка при обработке файла {}\nСтраница {}, ячейка {}{}'
                                 .format(filename, sheet.title, 'ABCDEFGHI'[i], j))
        datas.append(data)
    return datas


def read_xlsx_sample(filename):
    try:
        #для чтения xlsx
        wb = load_workbook(filename=filename)
    except Exception:
        raise ParseError("Невозможно открыть файл образца {}"
                         "\nВозможно он открыт в другой программе".format(filename))
    #инициализация листа с информацией о пациенте из xlsx
    sheet = wb.get_active_sheet()

    datas = []

    rows = sheet.max_row

    for i in range(2, rows + 1):
        data = []
        for j in range(1, 8 + 1):
            try:
                cell = sheet.cell(row=i, column=j)
                if sheet.cell(1, j).value == "date":
                    data.append(str(cell.value))
                else:
                    data.append(float(cell.value))
            except Exception:
                raise ParseError('Ошибка при обработке файла {}\nСтраница {}, ячейка {}{}'
                                 .format(filename, sheet.title, 'ABCDEFGHI'[i], j))
        datas.append(data)
    return datas

#какая-то шляпа, которая делает график из подсчитанных данных, возвращает буффер байтов
def plot_image(plot_func, *args, **kwargs):
    figure = Figure(dpi=100)
    canvas = FigureCanvasQT(figure)
    plot_func(*args, figure)

    buffer = BytesIO()
    canvas.print_figure(buffer)
    if kwargs.get('io', False):
        buffer.seek(0)
        return buffer
    else:
        return buffer.getvalue()
